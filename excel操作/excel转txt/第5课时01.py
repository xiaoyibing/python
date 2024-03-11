import openpyxl

sheets = ['598854', '598856', '598858']
txt_file_paths = ['598854课程.txt', '598856课程.txt', '598858课程.txt']

# 读取Excel文件
excel_file_path = '（前3次）2022级导论考试--成绩表.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)

def Write_txt(i):
    sheet = workbook[sheets[i]]
    txt_file_path = txt_file_paths[i]
    with open(txt_file_path, 'w') as txt_file:
        txt_file.truncate(0)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            # 仅选择列 A 到列 H 的数据
            row_data = ','.join(["  " if cell is None else str(cell) for cell in row[:8]])
            txt_file.write(row_data + '\n')

Write_txt(0)
Write_txt(1)
Write_txt(2)
workbook.close()
