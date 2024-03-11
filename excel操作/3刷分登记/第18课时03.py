'''任务3：刷分成绩登记（导论课）
已给定第01次的答卷及成绩数据文件，请提取其中的个人成绩，填写到任务1的《班级名单》文件中第
二列，提取规则为：多次刷分以个人最好成绩为准。若答题时间超过180秒，每超过2秒扣1分，比如
181-182秒扣1分，183-184秒扣1分...
'''
ls1=[]
a={}
from openpyxl import load_workbook
def wook1(aaa):
    wb = load_workbook(aaa)
    ws = wb['Sheet1']
    for row in ws.iter_rows(min_row=2, values_only=True):
        i, d, j = row[8], row[7], row[3]
        ls1.append([i, d, j[:-1]])  # 必须带【】
    wb.save(aaa)
wook1('导论考试01--进制转换1.xlsx')
wook1('导论考试01--进制转换2.xlsx')
wook1('导论考试01--进制转换3.xlsx')
wook1('导论考试01--进制转换4.xlsx')
wook1('导论考试01--进制转换5.xlsx')


ls2=[]
for i in ls1:
    if int(i[2])>180:
        dele = int((int(i[2])-180)/2)
        n=i[1]-1
        ls2.append([i[0],n])
    else:
        ls2.append([i[0],i[1]])

for i in ls2:
    if i[1] not in a:
        a[i[0]]=i[1]
for i in ls2:
    if  i[0] in a and i[1]>a[i[0]]:
        a[i[0]]=i[1]
print(a)
#-----------------------------------------
def wook2(bbb):
    wb = load_workbook('班级名单.xlsx')
    ws = wb[bbb]
    for row in range(1,ws.max_row+1):
        cell=ws.cell(row=row,column=3).value
        if cell in a:
            ws.cell(row=row,column=2,value=a[cell])
        else:
            ws.cell(row=row, column=2, value=0)
    wb.save('班级名单.xlsx')
wook2('导论682006')
wook2('导论655694')
wook2('导论655687')
wook2('导论655686')
wook2('导论重修群')
