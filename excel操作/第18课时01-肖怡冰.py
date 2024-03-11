'''任务1：昵称计分器（导论课）
有数据文件《QQ群昵称总表》和《班级名单》，内有4个头，一个重修补充名单，
对其QQ昵称进行评分，
标准为：1）学号姓名，100分；2）含有“学号”、“姓名”，90分；
3）只含有“姓名”无“学号”，80分；
4）无姓名的昵称或未加群，0分
输出：直接将得分添加了数据文件《班级名单》中第一列
'''
#------------QQ群昵称总表.xlsx------------------
ls=[]
from openpyxl import load_workbook
wb=load_workbook('QQ群昵称总表.xlsx')
ws=wb['导论862006']
for cell in ws['D']:
    ls.append(cell.value)
del ls[0]
wb.save('QQ群昵称总表.xlsx')
#print(ls)
#---------------班级名单.xlsx-------------------
from openpyxl import load_workbook
ll=[]
wb=load_workbook('班级名单.xlsx')
ws=wb['导论682006']
for cell in ws['C']:
    ll.append(cell.value)
del ll[0]
wb.save('班级名单.xlsx')
#print(ll)
#--------------给昵称打分-----------------------
ls1=[]
for i in ls:
    if i!=None:
        ls1.append(i)
#print(ls1)

def chinese(n):#第n个字符是汉字的字符串
    return 19968 <= ord(n) <= 40959
ls2=[item for item in ls1 if item.startswith('2023003') and len(item) >= 10 and chinese(item[10])]
#print(ls2)#100分
ls3=[item for item in ls1 if item.startswith('2023003') and item not in ls2]
#print(ls3)#含学号和姓名,但不规范（90）
ls4=[item for item in ls1 if '2023003' not in item]
#print(ls4)#不含有学号的80
#---------------将姓名与分数对应一一写入字典a------------
a={}#子函数用于填写字典a
def fen(x,y):
    for j in range(len(x)):
        if i in x[j]:
            a[i]=y
for i in ll:
    fen(ls2,100)
    fen(ls3,90)
    fen(ls4,80)
for i in ll:
    if i not in a:
        a[i]=0
print(a)
#--------------------填写Excel表---------------------
n=2
for key in ll:
    ws.cell(row=n,column=7,value=a[key])
    n+=1
wb.save('班级名单.xlsx')



