import turtle as t
import openpyxl as op

a={}
b = {}
c = []
d=[]
def x1():
    a.clear()
    xb1 = op.load_workbook('用户.xlsx')
    ws = xb1['Sheet1']
    for row in ws.rows:
        A = row[0].value
        B = row[1].value
        a[A] = B
    xb1.save('用户.xlsx')

    print(a)
def x2():
    b.clear()
    xb2 = op.load_workbook('商品.xlsx')
    ws = xb2['Sheet1']
    for row in ws.rows:
        A = row[0].value
        B = row[1].value
        C=row[2].value
        b[A] = [B,C]
    xb2.save('商品.xlsx')
    print(b)
def register():
    # -----用户管理--登录+注册系统-------
    while (1):
        x1()
        denglu = input('--登录请按--1\n--注册请按--2')
        if denglu == '1':
            xingming = input('请输入员工码：')
            mima = input('请输入密码：')
            if xingming in a and mima == a[xingming][0]:  # 一定是a[xingming][0]
                print('yes')
                break
            else:
                print('密码或员工码错误。')
        if denglu == '2':
            xingming = input('请输入员工码：')
            mima_1 = input('请输入密码：')
            mima_2 = input('请重复输入密码：')
            if mima_1 != mima_2:
                print('密码与重复密码不相同。')
            else:
                xb1 = op.load_workbook('用户.xlsx')
                ws = xb1['Sheet1']
                ws.append([xingming,mima_1])
                xb1.save('用户.xlsx')
def add():#增
    x2()
    name = input('请录入商品名称：')
    number = input("请录入商品条码号：")
    money = input('请录入商品单价：')
    if number not in b:
        xb2 = op.load_workbook('商品.xlsx')
        ws = xb2['Sheet1']
        ws.append([number, name,money])
        xb2.save('商品.xlsx')
    else:
        print('该商品已被录入')
def dele():#删
    x2()
    number = input("请录入商品条码号：")
    if number not in b:
        print('未找到该条码号')
    else:
        xb2 = op.load_workbook('商品.xlsx')
        ws = xb2['Sheet1']
        for i in range(ws.max_row,0,-1):
            if ws.cell(row=i,column=1).value==number:
                ws.delete_rows(i)
        xb2.save('商品.xlsx')
        print('已删除')
def re():#改
    x2()
    number1=input('请输入原来的条码号：')
    if number1 in b:
        number2 = input('请输入新的条码号：')
        name=input('请输入新的商品名：')
        money=input('请输入新的价格：')
        xb2 = op.load_workbook('商品.xlsx')
        ws = xb2['Sheet1']
        for i in range(1,ws.max_row+1,1):
            if ws.cell(row=i,column=1).value==number1:
                ws.delete_rows(i)
                ws.append([number2,name,money])
        xb2.save('商品.xlsx')
    else:
        print('该商品不存在')
def cha():#查
    x2()
    print('现在进行的是  商品扫码查询')
    number = input('请扫码：')
    if number in b:
        print(f"该商品的名称是：{b[number][0]}")
        print(f"该商品的单价是：{b[number][1]}")
    else:
        print('该商品不存在。')
def shouyin(): #收银
    c.clear()
    d.clear()
    x2()
    n='1'
    global dollar
    dollar=0
    while n=='1':
        numbers=input('请扫码：')
        if numbers in b:
            print(f'{b[numbers][0]}*1')
            dd=float(b[numbers][1])
            dollar+=dd
            c.append(b[numbers][0])
        else:
            print('该商品不存在。')
        n=input('----继续收银请按 1  \n-----结束收银请按  2')
    for i in c:
        if [i,c.count(i)] not in d:
            d.append([i,c.count(i)])
def draw1():
    gg = t.Screen()
    gg.title("小票")  # 名字
    gg.bgcolor('white')  # 颜色
    gg.setup(450, 700)  # 大小
def draw2():#小票
    draw1()
    t.hideturtle()
    for i in range(len(d)):
        t.penup()
        t.goto(-200,300-40*i)
        t.write(f'{d[i][0]}  *{d[i][1]}',font=('宋体',25,'normal'))
    t.goto(-200,300-40*(len(c)+1))
    t.write(f'商品总价为：{dollar}元',font=('宋体',25,'normal'))
#-----------------主程序--------
register()
while(1):
    q=input('增---1   \n  删-----2   \n改----3 \n 查-----4  \n 收银----5退出----6')
    if q=='1':
        add()
    if q=='2':
        dele()
    if q=='3':
        re()
    if q=='4':
        cha()
    if q=='5':
        shouyin()
        draw2()
    if q=='6':
        break