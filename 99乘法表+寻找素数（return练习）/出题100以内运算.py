import random
from openpyxl import load_workbook
yn=[]
name=[]
def jia():
    global dan1
    global grace
    while(1):
        a = random.randint(1, 100)
        b = random.randint(1, 100)
        if a + b <= 100 and dan1 >= 5:
            if a > 10 and b > 10:
                print(f'{a}+{b}=')
                answer = int(input('请输入你的运算结果：'))
                if answer == (a + b):
                    print('正确')
                    grace += 10
                else:
                    print(f'错误\n正确答案是：{a + b}')
                yn.append([f'{a}+{b}=', a + b, answer])

                break
        elif a + b <= 100 and dan1 < 5:
            print(f'{a}+{b}=')
            answer = int(input('请输入你的运算结果：'))
            if answer == (a + b):
                print('正确')
                grace += 10
            else:
                print(f'错误\n正确答案是：{a + b}')
            yn.append([f'{a}+{b}=', a + b, answer])
            if a < 10 or b < 10:
                dan1 += 1
            break

def jian():
    global dan1
    global grace

    while(1):
        a = random.randint(1, 100)
        b = random.randint(1, 100)
        if a - b >= 0 and dan1 >= 5:
            if a > 10 and b > 10:
                print(f'{a}-{b}=')
                answer = int(input('请输入你的运算结果：'))
                if answer == (a - b):
                    print('正确')
                    grace += 10
                else:
                    print(f'错误\n正确答案是：{a - b}')
                yn.append([f'{a}-{b}=', a - b, answer])


                break
        elif a - b >= 0 and dan1 < 5:
            print(f'{a}-{b}=')
            answer = int(input('请输入你的运算结果：'))
            if answer == (a - b):
                print('正确')
                grace += 10
            else:
                print(f'错误\n正确答案是：{a - b}')
            yn.append([f'{a}-{b}=', a - b, answer])


            if a < 10 or b < 10:
                dan1 += 1
            break
def cheng():
    global grace
    global dan1
    while(1):
        a = random.randint(1, 100)
        b = random.randint(1, 100)
        if  dan1 >= 5:
            a=10
            b=10
            print(f'{a}*{b}=')
            answer = int(input('请输入你的运算结果：'))
            if answer == (a * b):
                print('正确')
                grace += 10

            else:
                print(f'错误\n正确答案是：{a * b}')
            yn.append([f'{a}-{b}=', a * b, answer])
            break
        elif  a * b <= 100 and dan1 < 5:

            print(f'{a}*{b}=')
            answer = int(input('请输入你的运算结果：'))
            if answer == (a * b):
                print('正确')
                grace += 10
            else:
                print(f'错误\n正确答案是：{a * b}')
            yn.append([f'{a}*{b}=', a * b, answer])

            if a < 10 or b < 10:
                dan1 += 1
            break
def chu():
    global grace
    global dan1
    while(1):
        a = random.randint(1, 100)
        b = random.randint(1, 100)
        if a / b ==int(a/b) and dan1 >= 5:
            if a > 10 and b > 10:
                print(f'{a}/{b}=')
                answer = int(input('请输入你的运算结果：'))
                if answer == (a / b):
                    print('正确')
                    grace += 10
                else:
                    print(f'错误\n正确答案是：{a / b}')
                yn.append([f'{a}/{b}=', a / b, answer])
                break
        elif a / b ==int(a/b) and dan1 < 5:
            print(f'{a}/{b}=')
            answer = int(input('请输入你的运算结果：'))
            if answer == (a / b):
                print('正确')
                grace += 10
            else:
                print(f'错误\n正确答案是：{a / b}')
            yn.append([f'{a}/{b}=', a / b, answer])

            if a < 10 or b < 10:
                dan1 += 1
            break
question='y'
people=0
jjc=[]
while(question=='y'):
    na=input('请输入姓名：')
    people+=1
    jjcc1 = ['+', '-',  '/', '+','*', '-', '/']
    jjcc = ['+', '-', '*', '/']
    number = 1
    grace=0
    p1 = 0
    p2 = 0
    p3 = 0
    p4 = 0
    dan1 = 0
    cheng()
    number += 1
    p3 += 1
    while (number <= 8):
        jc = random.choice(jjcc1)
        jjcc1.remove(jc)
        if jc == '+':
            jia()
            number += 1
            p1+=1
        if jc == '-':
            jian()
            number += 1
            p2+=1
        if jc == '*':
            cheng()
            number += 1
            p3+=1
        if jc == '/':
            chu()
            number += 1
            p4+=1
    while (number <= 10):
        jc = random.choice(jjcc)
        if jc == '+':
            jia()
            number += 1
            p1+=1
        if jc == '-':
            jian()
            number += 1
            p2+=1
        if jc == '*':
            cheng()
            number += 1
            p3+=1
        if jc == '/':
            chu()
            number += 1
            p4+=1
    question=input('是否继续： y/n')
    name.append([na,grace])
    jjc.append([p1,p2,p3,p4,dan1])
wb=load_workbook('成绩表.xlsx')
ws=wb['Sheet1']
for i in range(len(name)):
    ws.cell(row=i+1,column=1,value=name[i][0])
    ws.cell(row=i+1,column=2,value=name[i][1])
wb.save('成绩表.xlsx')
wb=load_workbook('分析.xlsx')
ws=wb['Sheet1']
for i in range(10):
    ws.cell(row=1,column=i*3+1,value=f'第{i+1}题')
    ws.cell(row=1,column=i*3+2,value='正确结果')
    ws.cell(row=1,column=i*3+3,value='答题结果')
ws.cell(row=1,column=31,value='加法个数')
ws.cell(row=1,column=32,value='减法个数')
ws.cell(row=1,column=33,value='乘法个数')
ws.cell(row=1,column=34,value='除法个数')
ws.cell(row=1,column=35,value='全一位数的运算题数量')
for j in range(people):
    for i in range(10):
        ws.cell(row=j + 2, column=i * 3 + 1, value=yn[i+10*j][0])
        ws.cell(row=j + 2, column=i * 3 + 2, value=yn[i+10*j][1])
        ws.cell(row=j + 2, column=i * 3 + 3, value=yn[i+10*j][2])
    ws.cell(row=2 + j, column=31, value=jjc[j][0])
    ws.cell(row=2 + j, column=32, value=jjc[j][1])
    ws.cell(row=2 + j, column=33, value=jjc[j][2])
    ws.cell(row=2 + j, column=34, value=jjc[j][3])
    ws.cell(row=2 + j, column=35, value=jjc[j][4])



wb.save('分析.xlsx')


