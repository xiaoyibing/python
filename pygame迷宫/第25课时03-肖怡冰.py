import time
import pygame
import wall03

# ================================================
class Mysprite(pygame.sprite.Sprite):
    def __init__(self, color,x, y):
        super().__init__()
        self.image = pygame.Surface((20, 20))
        self.image.fill(color)
        self.rect = self.image.get_rect()
        self.rect.topleft = (x, y)
# ==================================================
class Player(pygame.sprite.Sprite):
    def __init__(self, color, x, y):
        radius = 5
        super().__init__()
        self.image = pygame.Surface((radius * 2, radius * 2), pygame.SRCALPHA)
        pygame.draw.circle(self.image, color, (radius, radius), radius)
        self.rect = self.image.get_rect(center=(x, y))

class Draws:
    def __init__(self,wall_color):
        global my_sprites
        global my_sprites1
        global player_ball
        global player_group
        my_sprites = pygame.sprite.Group()
        my_sprites1 = pygame.sprite.Group()
        # road1 = write_maze.road3  # 文字版迷宫

        for i in road3:
            a, b = i
            sprite = Mysprite(wall_color,a * 20, b * 20)
            my_sprites.add(sprite)
        sprite =Mysprite((0,255,0),(HEIGHT-1)*20,(WIDTH-2)*20)
        my_sprites1.add(sprite)

        my_sprites.draw(screen)
        my_sprites1.draw(screen)

        player_group = pygame.sprite.Group()  # 创建新的精灵组
        # 创建一个小球，并将小球添加到新的精灵组中
        player_ball = Player((255, 0, 0), 10, 30)
        player_group.add(player_ball)

        player_group.draw(screen)

# ==============================================================

class Game:
    def __init__(self,wall_color):
        global screen
        self.up_move = False
        self.down_move = False
        self.left_move = False
        self.right_move = False
        self.speed_circle = 2
        # =======================================
        print('程序有bug，一次只能按一个按键，如果同时按两个按键就会穿墙，暂时还没解决这个bug')
        # ===============================================
        WIN_WIDTH = WIDTH*20
        WIN_HEIGHT = HEIGHT*20
        # =================================================
        pygame.init()
        screen = pygame.display.set_mode((WIN_HEIGHT,WIN_WIDTH))
        pygame.display.set_caption('迷宫')
        screen.fill((225, 225, 225))
    #     ===========================
        Draws(wall_color)
        pygame.display.flip()  # 刷新
    def xunhuan(self):
        # global my_sprites
        # 使窗口一直运行
        while True:
            # ======================================================

            if (HEIGHT-1 )*20<= player_ball.rect.x <= HEIGHT*20 and (WIDTH-2 )*20 <= player_ball.rect.y <= (WIDTH-1 )*20:
                font = pygame.font.Font(None, 200)
                write_win = font.render('Win!', True, (225, 0, 0))
                screen.fill((225, 225, 225))
                screen.blit(write_win, (0, 0))
                pygame.display.update()
                time.sleep(2)
                return
            else:
                # ======================================================

                # 检测事件
                for event in pygame.event.get():
                    # ========================================键盘事件
                    if event.type == pygame.KEYDOWN:
                        if event.key == 1073741906:
                            self.down_move = True
                        if event.key == 1073741905:
                            self.up_move = True
                        if event.key == 1073741904:
                            self.left_move = True
                        if event.key == 1073741903:
                            self.right_move = True

                    if event.type == pygame.KEYUP:
                        if event.key == 1073741906:
                            self.down_move = False
                        if event.key == 1073741905:
                            self.up_move = False
                        if event.key == 1073741904:
                            self.left_move = False
                        if event.key == 1073741903:
                            self.right_move = False
                    # 检测关闭按钮
                    if event.type == pygame.QUIT:
                        exit()
                if self.up_move == True:
                    player_ball.rect.y += self.speed_circle
                if self.down_move == True:
                    player_ball.rect.y -= self.speed_circle
                if self.left_move == True:
                    player_ball.rect.x -= self.speed_circle
                if self.right_move == True:
                    player_ball.rect.x += self.speed_circle
                    # 更新并绘制所有精灵

                # ======================================================
                screen.fill((225, 225, 225))
                my_sprites.draw(screen)
                my_sprites1.draw(screen)
                player_group.update()
                player_group.draw(screen)

                # =======================================================
                # 碰撞检测

                collided_sprites = pygame.sprite.spritecollide(player_ball, my_sprites, False)
                if collided_sprites:
                    for collided_sprite in collided_sprites:
                        # 判断移动方向，并相应地调整玩家精灵的位置
                        if self.up_move:
                            player_ball.rect.bottom = collided_sprite.rect.top
                        elif self.down_move:
                            player_ball.rect.top = collided_sprite.rect.bottom
                        elif self.left_move:
                            player_ball.rect.left = collided_sprite.rect.right
                        elif self.right_move:
                            player_ball.rect.right = collided_sprite.rect.left
                else:
                    pass
                # ======================================================

                pygame.display.update()
                clock = pygame.time.Clock()
                clock.tick(60)  # 控制帧率为60fps

class  D:
    def __init__(self,nnn,wall_color):
        global road3
        global WIDTH
        global HEIGHT
        from openpyxl import load_workbook
        road3=[]
        wookbook=load_workbook(nnn)
        wooksheet=wookbook['Sheet1']
        HEIGHT= wooksheet.max_column
        WIDTH = wooksheet.max_row
        # 获取 单元格数据
        for row in wooksheet.iter_rows(min_row=1, max_row=wooksheet.max_row, min_col=1, max_col=wooksheet.max_column):
            for cell in row:
                if cell.value is not None:
                    # 将有内容的单元格的行和列信息添加到列表中
                    road3.append((cell.column-1, cell.row-1))
        game=Game(wall_color)
        game.xunhuan()


# 主程序==========================================
#D('墙03.xlsx',(0, 20, 120))
D('墙04.xlsx',(0, 20, 120))

D('墙05.xlsx',(210, 105, 30))

D('墙06.xlsx',(120, 20, 120))
